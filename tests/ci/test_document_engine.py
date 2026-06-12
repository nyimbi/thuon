# tests/ci/test_document_engine.py
"""Tests for core/document_engine.py and capabilities/document_generator.py"""
import sys, os, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '../../thuon_platform'))

from unittest.mock import MagicMock, patch
import pytest


# ── document_engine unit tests ───────────────────────────────────────────────

class TestDocxGeneration:
	def test_creates_file(self):
		from core.document_engine import _docx
		with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as f:
			path = f.name
		try:
			result = _docx('# Heading\n## Section\n- bullet\nParagraph text.', 'Test Doc', path)
			assert result == path
			assert os.path.getsize(path) > 0
		finally:
			os.unlink(path)

	def test_handles_empty_content(self):
		from core.document_engine import _docx
		with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as f:
			path = f.name
		try:
			_docx('', 'Empty', path)
			assert os.path.exists(path)
		finally:
			os.unlink(path)

	def test_all_heading_levels(self):
		from core.document_engine import _docx
		content = '# H1\n## H2\n### H3\n- bullet\nplain text'
		with tempfile.NamedTemporaryFile(suffix='.docx', delete=False) as f:
			path = f.name
		try:
			_docx(content, 'Headings', path)
			assert os.path.getsize(path) > 0
		finally:
			os.unlink(path)


class TestPdfGeneration:
	def test_creates_pdf_file(self):
		from core.document_engine import _pdf
		with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
			path = f.name
		try:
			result = _pdf('# Title\n## Section\nContent here.', 'Test PDF', path)
			assert result == path
			assert os.path.getsize(path) > 1000
			with open(path, 'rb') as pf:
				assert pf.read(4) == b'%PDF'
		finally:
			os.unlink(path)

	def test_bullet_points_in_pdf(self):
		from core.document_engine import _pdf
		with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
			path = f.name
		try:
			_pdf('- item one\n- item two\n* item three', 'Bullets', path)
			assert os.path.getsize(path) > 0
		finally:
			os.unlink(path)


class TestPptxGeneration:
	def test_creates_pptx_file(self):
		from core.document_engine import _pptx
		with tempfile.NamedTemporaryFile(suffix='.pptx', delete=False) as f:
			path = f.name
		try:
			result = _pptx('# Slide 1\nContent A\n# Slide 2\nContent B', 'Deck', path)
			assert result == path
			assert os.path.getsize(path) > 0
		finally:
			os.unlink(path)

	def test_explicit_slides(self):
		from core.document_engine import _pptx
		slides = [
			{'title': 'Intro', 'content': 'Welcome'},
			{'title': 'Data', 'content': 'Numbers here'},
		]
		with tempfile.NamedTemporaryFile(suffix='.pptx', delete=False) as f:
			path = f.name
		try:
			_pptx('', 'Explicit Slides', path, slides=slides)
			assert os.path.getsize(path) > 0
		finally:
			os.unlink(path)


class TestXlsxGeneration:
	def test_creates_xlsx_file(self):
		from core.document_engine import _xlsx
		rows = [
			{'Name': 'Alice', 'Revenue': 100000, 'Growth': '20%'},
			{'Name': 'Bob',   'Revenue': 85000,  'Growth': '15%'},
		]
		with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
			path = f.name
		try:
			result = _xlsx(rows, 'Q4 Data', path)
			assert result == path
			assert os.path.getsize(path) > 0
		finally:
			os.unlink(path)

	def test_empty_rows(self):
		from core.document_engine import _xlsx
		with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
			path = f.name
		try:
			_xlsx([], 'Empty Sheet', path)
			assert os.path.exists(path)
		finally:
			os.unlink(path)

	def test_sheet_name_truncated_at_31(self):
		from core.document_engine import _xlsx
		from openpyxl import load_workbook
		rows = [{'col': 'val'}]
		with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
			path = f.name
		try:
			_xlsx(rows, 'A' * 50, path)
			wb = load_workbook(path)
			assert len(wb.sheetnames[0]) <= 31
		finally:
			os.unlink(path)


class TestGenerateDocumentDispatch:
	def test_dispatches_to_docx(self):
		from core.document_engine import generate_document
		with tempfile.TemporaryDirectory() as d:
			path = os.path.join(d, 'out.docx')
			result = generate_document('docx', 'Hello world', 'Title', path)
			assert result == path
			assert os.path.exists(path)

	def test_dispatches_to_pdf(self):
		from core.document_engine import generate_document
		with tempfile.TemporaryDirectory() as d:
			path = os.path.join(d, 'out.pdf')
			result = generate_document('pdf', 'Hello world', 'Title', path)
			assert result == path

	def test_dispatches_to_xlsx_with_rows(self):
		from core.document_engine import generate_document
		with tempfile.TemporaryDirectory() as d:
			path = os.path.join(d, 'out.xlsx')
			rows = [{'A': 1, 'B': 2}]
			result = generate_document('xlsx', '', 'Title', path, rows=rows)
			assert result == path

	def test_unknown_format_raises(self):
		from core.document_engine import generate_document
		with tempfile.TemporaryDirectory() as d:
			with pytest.raises(ValueError, match='Unsupported format'):
				generate_document('txt', 'content', 'title', os.path.join(d, 'out.txt'))

	def test_creates_parent_dirs(self):
		from core.document_engine import generate_document
		with tempfile.TemporaryDirectory() as d:
			path = os.path.join(d, 'nested', 'dir', 'out.docx')
			generate_document('docx', 'content', 'title', path)
			assert os.path.exists(path)


# ── DocumentGenerator capability tests ──────────────────────────────────────

class TestDocumentGeneratorCapability:
	def _make_gen(self, ai_response='# Report Title\nContent here.'):
		mock_ai = MagicMock()
		mock_ai.generate_text.return_value = ai_response
		from capabilities.document_generator import DocumentGenerator
		with tempfile.TemporaryDirectory() as d:
			gen = DocumentGenerator(mock_ai, output_dir=d)
			return gen, d

	def test_generate_docx_returns_dict(self):
		mock_ai = MagicMock()
		mock_ai.generate_text.return_value = '# Q4 Report\n## Results\nRevenue up 20%.'
		from capabilities.document_generator import DocumentGenerator
		with tempfile.TemporaryDirectory() as d:
			gen = DocumentGenerator(mock_ai, output_dir=d)
			result = gen.generate('Q4 results', format='docx')
			assert result['status'] == 'ok'
			assert result['format'] == 'docx'
			assert os.path.exists(result['output_path'])

	def test_generate_pdf(self):
		mock_ai = MagicMock()
		mock_ai.generate_text.return_value = '# Annual Report\n## Summary\nGood year.'
		from capabilities.document_generator import DocumentGenerator
		with tempfile.TemporaryDirectory() as d:
			gen = DocumentGenerator(mock_ai, output_dir=d)
			result = gen.generate('Annual report', format='pdf')
			assert result['status'] == 'ok'
			assert result['format'] == 'pdf'
			assert os.path.exists(result['output_path'])

	def test_generate_pptx(self):
		mock_ai = MagicMock()
		mock_ai.generate_text.return_value = '# Slide 1\nIntro content.\n# Slide 2\nData content.'
		from capabilities.document_generator import DocumentGenerator
		with tempfile.TemporaryDirectory() as d:
			gen = DocumentGenerator(mock_ai, output_dir=d)
			result = gen.generate('Q4 presentation', format='pptx')
			assert result['status'] == 'ok'
			assert os.path.exists(result['output_path'])

	def test_generate_xlsx_with_rows(self):
		mock_ai = MagicMock()
		from capabilities.document_generator import DocumentGenerator
		rows = [{'Product': 'A', 'Sales': 1000}, {'Product': 'B', 'Sales': 2000}]
		with tempfile.TemporaryDirectory() as d:
			gen = DocumentGenerator(mock_ai, output_dir=d)
			result = gen.generate('Sales data', format='xlsx', rows=rows)
			assert result['status'] == 'ok'
			assert result['format'] == 'xlsx'
			assert os.path.exists(result['output_path'])

	def test_custom_output_path(self):
		mock_ai = MagicMock()
		mock_ai.generate_text.return_value = '# Title\nContent.'
		from capabilities.document_generator import DocumentGenerator
		with tempfile.TemporaryDirectory() as d:
			custom = os.path.join(d, 'custom_name.docx')
			gen = DocumentGenerator(mock_ai, output_dir=d)
			result = gen.generate('topic', format='docx', output_path=custom)
			assert result['output_path'] == custom

	def test_extract_title_from_heading(self):
		mock_ai = MagicMock()
		from capabilities.document_generator import DocumentGenerator
		gen = DocumentGenerator(mock_ai)
		title = gen._extract_title('# The Real Title\nsome content', 'fallback')
		assert title == 'The Real Title'

	def test_extract_title_fallback(self):
		mock_ai = MagicMock()
		from capabilities.document_generator import DocumentGenerator
		gen = DocumentGenerator(mock_ai)
		title = gen._extract_title('', 'fallback topic')
		assert title == 'fallback topic'

	def test_extract_slides_splits_on_headings(self):
		mock_ai = MagicMock()
		from capabilities.document_generator import DocumentGenerator
		gen = DocumentGenerator(mock_ai)
		content = '# Intro\nHello world.\n# Data\nNumbers here.'
		slides = gen._extract_slides(content)
		assert len(slides) == 2
		assert slides[0]['title'] == 'Intro'
		assert 'Hello world' in slides[0]['content']

	def test_word_count_in_result(self):
		mock_ai = MagicMock()
		mock_ai.generate_text.return_value = ' '.join(['word'] * 100)
		from capabilities.document_generator import DocumentGenerator
		with tempfile.TemporaryDirectory() as d:
			gen = DocumentGenerator(mock_ai, output_dir=d)
			result = gen.generate('topic', format='docx')
		assert result['word_count'] == 100
