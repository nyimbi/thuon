import csv
import os
from typing import Any

try:
	import openpyxl
except ImportError:
	openpyxl = None


class ExcelReader:

	def read(self, file_path: str, sheet_name: str = '', max_rows: int = 500) -> dict[str, Any]:
		try:
			if not os.path.exists(file_path):
				return {'status': 'error', 'error': f'File not found: {file_path}'}

			ext = os.path.splitext(file_path)[1].lower()

			if ext == '.csv':
				return self._read_csv(file_path, max_rows)
			elif ext == '.xlsx':
				return self._read_xlsx(file_path, sheet_name, max_rows)
			else:
				return {'status': 'error', 'error': f'Unsupported file type: {ext}. Supported: .xlsx, .csv'}
		except Exception as e:
			return {'status': 'error', 'error': str(e)}

	def _read_csv(self, file_path: str, max_rows: int) -> dict[str, Any]:
		try:
			with open(file_path, newline='', encoding='utf-8-sig') as f:
				reader = csv.DictReader(f)
				headers = list(reader.fieldnames or [])
				rows: list[dict] = []
				truncated = False
				for i, row in enumerate(reader):
					if i >= max_rows:
						truncated = True
						break
					rows.append(dict(row))

			sheet = {
				'name': os.path.basename(file_path),
				'headers': headers,
				'rows': rows,
				'row_count': len(rows),
				'col_count': len(headers),
			}
			return {
				'status': 'success',
				'file_path': file_path,
				'sheets': [sheet],
				'sheet_count': 1,
				'truncated': truncated,
			}
		except Exception as e:
			return {'status': 'error', 'error': str(e)}

	def _read_xlsx(self, file_path: str, sheet_name: str, max_rows: int) -> dict[str, Any]:
		try:
			if openpyxl is None:
				return {'status': 'error', 'error': 'Package openpyxl not installed. Run: uv add openpyxl'}

			wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
			sheet_names = [sheet_name] if sheet_name else wb.sheetnames
			sheets: list[dict] = []
			truncated = False

			for name in sheet_names:
				if name not in wb.sheetnames:
					return {'status': 'error', 'error': f'Sheet not found: {name}'}

				ws = wb[name]
				rows_iter = ws.iter_rows(values_only=True)

				# first row is headers
				try:
					header_row = next(rows_iter)
				except StopIteration:
					sheets.append({'name': name, 'headers': [], 'rows': [], 'row_count': 0, 'col_count': 0})
					continue

				headers = [str(h) if h is not None else '' for h in header_row]
				rows: list[dict] = []
				sheet_truncated = False

				for i, row in enumerate(rows_iter):
					if i >= max_rows:
						sheet_truncated = True
						truncated = True
						break
					rows.append(dict(zip(headers, [str(v) if v is not None else '' for v in row])))

				sheets.append({
					'name': name,
					'headers': headers,
					'rows': rows,
					'row_count': len(rows),
					'col_count': len(headers),
				})

			wb.close()
			return {
				'status': 'success',
				'file_path': file_path,
				'sheets': sheets,
				'sheet_count': len(sheets),
				'truncated': truncated,
			}
		except Exception as e:
			return {'status': 'error', 'error': str(e)}
